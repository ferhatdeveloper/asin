# -*- coding: utf-8 -*-
"""yemekcom-tarif-urunler-ExcelModule.xlsx -> TR/EN/AR çok dilli şablon üretir."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SRC = Path(__file__).resolve().parent / "yemekcom-tarif-urunler-ExcelModule.xlsx"
DST = Path(__file__).resolve().parent / "yemekcom-tarif-urunler-ExcelModule_TR_EN_AR.xlsx"

# Kaynak sütun indeksleri (1-based, orijinal "Ürünler" sayfası)
COL = {
    "urun_kodu": 1,
    "urun_adi": 2,
    "barkod": 3,
    "kategori": 4,
    "grup_kodu": 5,
    "marka": 6,
    "birim": 7,
    "alis": 8,
    "satis": 9,
    "kdv": 10,
    "min_stok": 11,
    "max_stok": 12,
    "ozel1": 13,
    "ozel2": 14,
    "ozel3": 15,
    "aciklama": 16,
    "gorsel": 17,
    "aktif": 18,
}

HEADERS = [
    "Ürün Kodu*",
    "Barkod",
    "Grup Kodu",
    "Marka",
    "Birim",
    "Alış Fiyatı",
    "Satış Fiyatı*",
    "KDV Oranı (%)",
    "Min Stok",
    "Max Stok",
    "Özel Kod 1",
    "Özel Kod 2",
    "Özel Kod 3",
    "Görsel URL",
    "Aktif (E/H)",
    "Ürün Adı (TR)*",
    "Ürün Adı (EN)",
    "Ürün Adı (AR)",
    "Kategori (TR)",
    "Kategori (EN)",
    "Kategori (AR)",
    "Açıklama (TR)",
    "Açıklama (EN)",
    "Açıklama (AR)",
]


def main() -> None:
    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    ws_src = wb_src[wb_src.sheetnames[0]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urunler_TR_EN_AR"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_wrap = Alignment(wrap_text=True, vertical="top")

    for c, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    max_r = ws_src.max_row
    for r in range(2, max_r + 1):
        out = 1
        ws.cell(r, out, ws_src.cell(r, COL["urun_kodu"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["barkod"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["grup_kodu"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["marka"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["birim"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["alis"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["satis"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["kdv"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["min_stok"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["max_stok"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["ozel1"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["ozel2"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["ozel3"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["gorsel"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["aktif"]).value)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["urun_adi"]).value)
        out += 1
        ws.cell(r, out, None)
        out += 1
        ws.cell(r, out, None)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["kategori"]).value)
        out += 1
        ws.cell(r, out, None)
        out += 1
        ws.cell(r, out, None)
        out += 1
        ws.cell(r, out, ws_src.cell(r, COL["aciklama"]).value)
        out += 1
        ws.cell(r, out, None)
        out += 1
        ws.cell(r, out, None)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = thin_wrap

    ws.freeze_panes = "A2"

    tab = Table(
        displayName="UrunlerI18n",
        ref=f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}",
    )
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # Arapça metin sütunları: hücre içi RTL (Excel içe/dışa aktarımda daha tutarlı)
    ar_cols = {18, 21, 24}  # Ürün Adı (AR), Kategori (AR), Açıklama (AR)
    for r in range(2, ws.max_row + 1):
        for c in ar_cols:
            cell = ws.cell(r, c)
            if cell.value is not None and str(cell.value).strip():
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    readingOrder=2,
                )

    ws_help = wb.create_sheet("Yardim_TR_EN_AR", 1)
    help_lines = [
        ("Amaç", "Bu dosya, ürün satırlarını tek tabloda TR/EN/AR dilleriyle doldurmanız için hazırlanmıştır."),
        ("Purpose", "This workbook lets you maintain product rows in one table with Turkish, English, and Arabic text fields."),
        ("الغرض", "يتيح لك هذا الملف الاحتفاظ بصفوف المنتجات في جدول واحد مع حقول نصية بالتركية والإنجليزية والعربية."),
        ("", ""),
        ("Doldurma", "TR sütunları kaynaktan kopyalanmıştır. EN ve AR sütunlarını çevirerek doldurun; kod ve fiyat sütunlarını dil bazında çoğaltmayın."),
        ("Filling", "TR columns are copied from the source. Fill EN and AR with translations; keep codes and prices language-neutral."),
        ("التعبئة", "أُخذت أعمدة TR من المصدر. املأ EN و AR بالترجمة؛ احتفظ بالرموز والأسعار دون تكرار لكل لغة."),
        ("", ""),
        ("İçe aktarma", "Başlık satırı sabittir. CSV/ETL kullanıyorsanız sütun adlarını anahtar olarak kullanın."),
        ("Import", "Keep the header row fixed. For CSV/ETL, use column names as stable keys."),
        ("الاستيراد", "أبقِ صف العناوين ثابتًا. لاستيراد CSV/ETL استخدم أسماء الأعمدة كمفاتيح ثابتة."),
    ]
    for i, (a, b) in enumerate(help_lines, start=1):
        ws_help.cell(i, 1, value=a)
        ws_help.cell(i, 2, value=b)
    ws_help.column_dimensions["A"].width = 22
    ws_help.column_dimensions["B"].width = 88

    widths = {
        1: 14,
        2: 16,
        3: 12,
        4: 10,
        5: 8,
        6: 10,
        7: 12,
        8: 8,
        9: 8,
        10: 8,
        11: 12,
        12: 12,
        13: 12,
        14: 48,
        15: 12,
        16: 28,
        17: 28,
        18: 28,
        19: 18,
        20: 18,
        21: 18,
        22: 44,
        23: 44,
        24: 44,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    wb_src.close()
    wb.save(DST)
    print("Wrote", DST)


if __name__ == "__main__":
    main()
